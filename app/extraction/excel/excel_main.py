# excel_main.py
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def excel_to_dataframe(excel_path: Path) -> pd.DataFrame:
    wb = load_workbook(excel_path, data_only=True)
    all_sheets_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = list(ws.iter_rows(min_row=1, values_only=True))

        if len(data) < 2:
            continue

        header_row = 0
        headers = list(data[header_row])

        records = []
        for row_idx in range(header_row + 1, len(data)):
            row_values = data[row_idx]
            if all(v is None for v in row_values):
                continue

            record = {}
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                record[header if header else f"Unnamed_{col_letter}"] = row_values[col_idx]
            records.append(record)

        if records:
            df = pd.DataFrame(records)
            all_sheets_data.append(df)

    wb.close()

    if all_sheets_data:
        return pd.concat(all_sheets_data, ignore_index=True)
    else:
        return pd.DataFrame()


# ====================== 修改重點 ======================

def main_excel(EXCEL_PATHS):
    """接收 Excel 路徑列表，回傳合併後的 DataFrame"""
    all_dfs = []
    for excel_path in EXCEL_PATHS:
        if not Path(excel_path).exists():
            print(f"⚠️ 找不到 Excel 檔案：{excel_path}")
            continue
        df = excel_to_dataframe(Path(excel_path))
        if not df.empty:
            all_dfs.append(df)
            print(f"✅ 已讀取 Excel：{excel_path}，共 {len(df)} 筆")

    if all_dfs:
        return pd.concat(all_dfs, ignore_index=True)
    else:
        return pd.DataFrame()
