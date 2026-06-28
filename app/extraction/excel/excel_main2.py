import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def excel_to_dataframe(excel_path: Path) -> pd.DataFrame:
    wb = load_workbook(excel_path, data_only=True)
    print(f"=== 檔案：{excel_path.name} ===")
    print(f"工作表列表：{wb.sheetnames}")

    all_sheets_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = list(ws.iter_rows(min_row=1, values_only=True))

        # 過濾完全空白的列
        data = [row for row in data if not all(v is None for v in row)]

        print(f"\n工作表 [{sheet_name}]：原始資料列數 = {len(data)}")

        if len(data) == 0:
            print("→ 此工作表完全空白，跳過")
            continue

        # 取第一行當表頭
        headers = list(data[0])
        headers = [
            h if h else f"Unnamed_{get_column_letter(col_idx + 1)}"
            for col_idx, h in enumerate(headers)
        ]
        print(f"→ 表頭：{headers}")

        records = []
        for row_values in data[1:]:
            if all(v is None for v in row_values):
                continue
            record = {headers[col_idx]: row_values[col_idx] for col_idx in range(len(headers))}
            records.append(record)

        if records:
            df = pd.DataFrame(records)
        else:
            df = pd.DataFrame(columns=headers)

        print(f"→ 讀取到 {len(df)} 筆資料")
        all_sheets_data.append(df)

    wb.close()

    if all_sheets_data:
        return pd.concat(all_sheets_data, ignore_index=True)
    else:
        return pd.DataFrame()


def main_excel(EXCEL_PATHS):
    if isinstance(EXCEL_PATHS, (str, Path)):
        EXCEL_PATHS = [EXCEL_PATHS]

    all_dfs = []
    for excel_path in EXCEL_PATHS:
        excel_path = Path(excel_path)
        if not excel_path.exists():
            print(f"⚠️ 找不到檔案：{excel_path}")
            continue
        df = excel_to_dataframe(excel_path)
        if not df.empty or len(df.columns) > 0:
            all_dfs.append(df)

    if all_dfs:
        return pd.concat(all_dfs, ignore_index=True)
    else:
        return pd.DataFrame()


if __name__ == "__main__":
    EXCEL_PATHS = ["/Users/mac/python project/invomag/app/extraction/excel/input/task2_700.xlsx"]
    df = main_excel(EXCEL_PATHS)
    print("\n=== 最終結果 ===")
    print(df)
    print(f"欄位名稱：{list(df.columns)}")
    print(f"資料筆數：{len(df)}")
