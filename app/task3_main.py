# task3_main.py
# Task3 = Task1 風格的結算報告 PDF(A) + Task2 風格的 Accumulator 合約 PDF(B) → Excel(B)
#
# 分兩部分：
#   Part A：沿用 Task2 既有流程，把 PDF(B) 的新合約資料寫進 Excel(B)（新增列）。
#           此時「KO Date / Expiry Date」欄位填的是合約「到期日」（expiration_date），
#           還不知道這筆合約是否真的提前 KO。
#   Part B：解析 PDF(A) 的「AQ/DQ KO:」區段，依 (AQ/DQ, Strike Px) 兩欄完全比對找出
#           Excel 既有列（# Shares Traded 在 Part A 階段還是空的，KO 發生前無法拿來比對，
#           所以不放進比對鍵）。若該列 Remarks 還是空的，代表這是第一次發現它已 KO，
#           這時才把：
#             - Remarks 填成字面值 "KO"
#             - # Shares Traded 填入 PDF(A) 的股數
#             - KO Date / Expiry Date 改填 PDF(A) 的真實 KO 日期（覆蓋掉 Part A 寫的合約到期日）
#           PDF(A) 的其他欄位（Ticker、total_amount...）不會拿來更新任何欄位。
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

from app.task1_main3 import main as run_task2_flow, OUTPUT_DIR, _to_excel_date
from app.extraction.pdf.pdf_to_json import extract_pdf_to_json
from app.extraction.pdf.json_to_df import process_aq_dq_ko

# Excel(B) 表頭文字 -> PDF(A) KO 記錄欄位。用來「找出對應列」，不是用來覆寫資料。
_MATCH_FIELDS = {
    "AQ / DQ": "type",
    "Strike Px": "price",
}
_REMARKS_COLUMN = "Remarks"
_SHARES_TRADED_COLUMN = "# Shares Traded"
_KO_DATE_COLUMN = "KO Date / Expiry Date"


def _find_header(ws):
    """回傳 (表頭列號, 欄名 -> 欄位編號)，欄名做 strip() 避免表頭文字尾端有空白/換行造成比對失敗。"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if any(v is not None for v in row):
            col_index = {
                str(name).strip(): idx + 1
                for idx, name in enumerate(row) if name
            }
            return row_idx, col_index
    return None, {}


def _norm_num(value):
    """把數字（int/float/字串）正規化成可比對的浮點數，None 則保留 None。"""
    if value is None:
        return None
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return value


def _set_date_cell(cell, date_str):
    value = _to_excel_date(date_str)
    cell.value = value
    if isinstance(value, datetime):
        cell.number_format = "yyyy/mm/dd"


def _apply_ko_updates(excel_path: Path, ko_df) -> bool:
    """掃描 Excel 既有列，依 (AQ/DQ, Strike Px) 找出對應列。
    若該列 Remarks 還是空的，代表第一次發現它已 KO：
        Remarks <- "KO"
        # Shares Traded <- PDF(A) 的股數
        KO Date / Expiry Date <- PDF(A) 的真實 KO 日期（覆蓋掉合約到期日）
    Remarks 已有值（包含先前已寫過 "KO"）一律 Pass，避免重複處理。
    回傳是否有任何更新。
    """
    if ko_df.empty:
        print("PDF(A) 沒有解析到 KO 記錄，略過 KO 回填")
        return False

    # AQ/DQ、Strike Px 欄位可能是公式，用 data_only=True 開一份「唯讀比對用」的副本讀取
    # 計算後的數值；實際寫入仍用 data_only=False 的版本，避免把其他欄位的公式洗掉。
    wb_values = load_workbook(excel_path, data_only=True)
    ws_values = wb_values.active

    wb = load_workbook(excel_path)
    ws = wb.active

    header_row_idx, col_index = _find_header(ws)
    required = list(_MATCH_FIELDS.keys()) + [_REMARKS_COLUMN, _SHARES_TRADED_COLUMN, _KO_DATE_COLUMN]
    missing = [c for c in required if c not in col_index]
    if header_row_idx is None or missing:
        print(f"⚠️ Excel 缺少回填 KO 資訊所需的欄位，略過：{missing}")
        wb.close()
        wb_values.close()
        return False

    # 用 (type, price) 建立 KO 對照表：type/price -> (#, ko_date)
    ko_lookup = {}
    for _, r in ko_df.iterrows():
        key = (r.get("type"), _norm_num(r.get("price")))
        ko_lookup[key] = (r.get("#"), r.get("ko_date"))

    aq_col = col_index["AQ / DQ"]
    strike_col = col_index["Strike Px"]
    remarks_col = col_index[_REMARKS_COLUMN]
    shares_col = col_index[_SHARES_TRADED_COLUMN]
    ko_date_col = col_index[_KO_DATE_COLUMN]

    updated = False
    value_rows = ws_values.iter_rows(min_row=header_row_idx + 1)
    write_rows = ws.iter_rows(min_row=header_row_idx + 1)
    for value_row, row in zip(value_rows, write_rows):
        aq_dq_val = value_row[aq_col - 1].value
        strike_val = value_row[strike_col - 1].value
        if aq_dq_val is None and strike_val is None:
            continue  # 整列空白

        key = (aq_dq_val, _norm_num(strike_val))
        remarks_cell = row[remarks_col - 1]

        # 規則：找得到對應的 KO 記錄，且 Excel Remarks 目前是空的，才處理；
        # Remarks 已有資料（包含已標過 "KO"）一律 Pass，避免重複覆寫。
        if key in ko_lookup and remarks_cell.value in (None, "", "Nil"):
            shares_val_pdf, ko_date_val = ko_lookup[key]
            remarks_cell.value = "KO"
            row[shares_col - 1].value = shares_val_pdf
            _set_date_cell(row[ko_date_col - 1], ko_date_val)
            updated = True

    if updated:
        wb.save(excel_path)
    wb.close()
    wb_values.close()
    return updated


def main(task):
    """task 需包含 pdf_path_a（結算報告 PDF）、pdf_path_b（Accumulator 合約 PDF）、excel。"""
    print("=== Task3：PDF(A)+PDF(B) → Excel(B) ===\n")

    # ---------- Part A：沿用 Task2 流程，把 PDF(B) 的新合約寫進 Excel ----------
    task2_task = dict(task)
    # task_name="task2" 只是為了讓 PDF 解析器/欄位對應走 Task2 分流（get_pdf_parser_for_task /
    # get_pdf_column_mapping 都是依 task_name 前綴判斷），輸出檔名要用 output_prefix 另外指定，
    # 否則 task1_main3.main() 產生檔名時會直接讀到 task_name="task2"，產生 "task2_Updated_..."。
    task2_task["output_prefix"] = task.get("task_name", "task3")
    task2_task["task_name"] = "task2"
    task2_task["pdf_path"] = task["pdf_path_b"]
    task2_task["pdf_to_json"] = task.get("pdf_to_json_b", task.get("pdf_to_json"))

    result_path = run_task2_flow(task2_task)
    excel_path = Path(result_path) if result_path else Path(task["excel"])

    # ---------- Part B：用 PDF(A) 的 KO 資訊回填 Remarks / # Shares Traded / KO Date ----------
    json_path_a = task.get("pdf_to_json_a", task.get("pdf_to_json"))
    pdf_a_data = extract_pdf_to_json(str(task["pdf_path_a"]), str(json_path_a))
    ko_df = process_aq_dq_ko(pdf_a_data)

    ko_updated = _apply_ko_updates(excel_path, ko_df)

    if not result_path and not ko_updated:
        print("✅ PDF(A)/(B) 與 Excel 完全一致，無需更新。")
        return None

    if ko_updated and not result_path:
        # Part A 沒有新增列，但 Part B 直接修改了原始 Excel，產生一份帶時間戳記的副本回傳，
        # 避免覆寫使用者上傳的原始檔案。
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        task_name = task.get("task_name", "task3")
        new_path = OUTPUT_DIR / f"{task_name}_Updated_{timestamp}.xlsx"
        wb = load_workbook(excel_path)
        wb.save(new_path)
        wb.close()
        return str(new_path)

    return str(excel_path)
