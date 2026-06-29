# task1_main3.py
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

from app.extraction.pdf.pdf_main import main_pdf
from app.extraction.excel.excel_main import main_excel
from app.mapping.mapping_main import compare_pdf_excel
from app.mapping.column_mapping import get_pdf_column_mapping

PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = PROJECT_ROOT / "final_output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

_PDF_DATE_FORMATS = ["%d-%b-%Y", "%d-%b-%y"]


def _to_excel_date(value):
    """把 PDF 解析出來的日期字串（如 16-Mar-2026）轉成 datetime，讓 Excel 以日期格式顯示。"""
    if not isinstance(value, str):
        return value
    for fmt in _PDF_DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return value


def main(task):
    print("正在使用 mapping_main.py 進行比對...\n")

    excel_path = Path(task["excel"])
    pdf_path = Path(task["pdf_path"])
    json_output_path = task["pdf_to_json"]

    # 讀取 PDF 與 Excel
    pdf_df = main_pdf(str(pdf_path), str(json_output_path), task)
    excel_df = main_excel([str(excel_path)])

    # ====================== 修改後的空值檢查 ======================
    if pdf_df.empty:
        print("⚠️ PDF 或 Excel 資料為空，無法進行比對")
        return None

    if excel_df.empty:
        print("⚠️ Excel 為空，但 PDF 有資料，將繼續新增")

    # 比對
    missing_keys, pdf_df = compare_pdf_excel(pdf_df, excel_df, task)

    print(f"\n需要新增的資料筆數：{len(missing_keys)}")

    # ====================== 修改重點 ======================
    if not missing_keys:
        print("✅ Excel 已包含 PDF 所有資料，無需更新。")
        return None   # [修改] 沒有差異就回傳 None

    # 取出要新增的資料（只取這個 task 有對應到 PDF 資料的欄位，其餘 Excel 欄位維持原值不動）
    write_columns = list(get_pdf_column_mapping(task).keys())
    to_add = pdf_df[pdf_df["_match_key"].isin(missing_keys)][write_columns].drop_duplicates().copy()

    print(f"\n發現 {len(to_add)} 筆需要新增的資料：")
    print(to_add)

    # ====================== 只有需要更新時才產生新檔案 ======================
    wb = load_workbook(excel_path)
    ws = wb.active

    # 依「表頭文字」找出每個欄位實際所在的欄位編號，而非假設欄位順序與 write_columns 一致。
    # 表頭不一定在第 1 列（例如 task2_700.xlsx 第 1 列是空白列，表頭在第 2 列），
    # 所以動態找出第一個「非全空白」的列當表頭，邏輯與 excel_main.py 的讀取方式一致。
    header_row = next(
        (row for row in ws.iter_rows(min_row=1, values_only=True) if any(v is not None for v in row)),
        ()
    )
    col_index_by_name = {name: idx + 1 for idx, name in enumerate(header_row) if name}

    missing_in_excel = [c for c in write_columns if c not in col_index_by_name]
    if missing_in_excel:
        print(f"⚠️ Excel 表頭找不到以下欄位，將略過寫入：{missing_in_excel}")

    start_row = ws.max_row + 1
    for r_idx, row in enumerate(to_add.itertuples(index=False), start_row):
        for col_name, value in zip(write_columns, row):
            if col_name in col_index_by_name:
                if "Date" in col_name:
                    value = _to_excel_date(value)
                cell = ws.cell(row=r_idx, column=col_index_by_name[col_name], value=value)
                if "Date" in col_name and isinstance(value, datetime):
                    cell.number_format = "yyyy/mm/dd"

    # 產生新的檔案名稱
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    task_name = task.get("task_name", "task")
    new_filename = f"{task_name}_Updated_{timestamp}.xlsx"
    new_file_path = OUTPUT_DIR / new_filename

    wb.save(new_file_path)
    wb.close()

    print(f"\n✅ 已產生更新後的 Excel：{new_file_path}")
    return str(new_file_path)   # [修改] 回傳新檔案路徑
